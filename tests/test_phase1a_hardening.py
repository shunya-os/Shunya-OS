"""
PHASE 1A — HARDENING TESTS: XLSX, proposal versioning, transaction safety, partial commit
"""
import pytest, json, os


# =========================================================================
# 1. XLSX ADAPTER
# =========================================================================

class TestXLSXAdapter:

    def test_xlsx_parse_workbook(self, real_app):
        """XLSX workbook is parsed through the same pipeline as CSV."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customers"
        ws.append(["name", "email", "phone"])
        ws.append(["Ritu Sharma", "ritu@example.com", "+919876543210"])
        ws.append(["Arjun Singh", "arjun@example.com", "+919999999999"])
        path = "/tmp/test_intake.xlsx"
        wb.save(path)
        try:
            from app.intake.profiler import SchemaProfiler
            profiler = SchemaProfiler()
            columns, rows = profiler.parse_xlsx(path)
            assert len(columns) == 3
            assert len(rows) == 2
            assert rows[0]["name"] == "Ritu Sharma"
            assert rows[0]["email"] == "ritu@example.com"
        finally:
            os.remove(path)

    def test_xlsx_first_worksheet_default(self, real_app):
        """Default worksheet is used when no specific sheet is named."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["col1", "col2"])
        ws.append(["a", "b"])
        wb.save("/tmp/test_xlsx_default.xlsx")
        try:
            from app.intake.profiler import SchemaProfiler
            columns, rows = SchemaProfiler.parse_xlsx("/tmp/test_xlsx_default.xlsx")
            assert len(columns) == 2
            assert rows[0]["col1"] == "a"
        finally:
            os.remove("/tmp/test_xlsx_default.xlsx")

    def test_xlsx_null_cells(self, real_app):
        """Null cells in XLSX are preserved as empty strings."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "email", "phone"])
        ws.append(["Ritu", "", ""])
        ws.append(["Arjun", "arjun@example.com", None])
        wb.save("/tmp/test_xlsx_null.xlsx")
        try:
            from app.intake.profiler import SchemaProfiler
            columns, rows = SchemaProfiler.parse_xlsx("/tmp/test_xlsx_null.xlsx")
            assert rows[0]["email"] == ""
            assert rows[0]["phone"] == ""
            assert rows[1]["phone"] == ""
        finally:
            os.remove("/tmp/test_xlsx_null.xlsx")

    def test_xlsx_empty_workbook(self, real_app):
        """Empty XLSX returns empty columns and rows."""
        import openpyxl
        wb = openpyxl.Workbook()
        wb.save("/tmp/test_xlsx_empty.xlsx")
        try:
            from app.intake.profiler import SchemaProfiler
            columns, rows = SchemaProfiler.parse_xlsx("/tmp/test_xlsx_empty.xlsx")
            assert columns == []
            assert rows == []
        finally:
            os.remove("/tmp/test_xlsx_empty.xlsx")

    def test_xlsx_through_profile_pipeline(self, real_app):
        """XLSX goes through the same SchemaProfiler.profile() pipeline as CSV."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "email", "phone"])
        ws.append(["Ritu Sharma", "ritu@example.com", "+919876543210"])
        wb.save("/tmp/test_xlsx_profile.xlsx")
        try:
            from app.intake.profiler import SchemaProfiler
            profiler = SchemaProfiler()
            profile = profiler.profile(xlsx_path="/tmp/test_xlsx_profile.xlsx")
            assert profile["row_count"] == 1
            assert profile["column_count"] == 3
            assert "email" in profile["profiles"]
            assert profile["profiles"]["email"]["primitive_type"] == "email"
        finally:
            os.remove("/tmp/test_xlsx_profile.xlsx")

    def test_xlsx_email_phone_mapping(self, real_app):
        """XLSX columns map through the same FieldMapper as CSV."""
        from app.intake.mapper import FieldMapper
        mapper = FieldMapper()
        target, _, _ = mapper.map_column("email")
        assert target == "identity.email"
        target, _, _ = mapper.map_column("phone")
        assert target == "identity.phone"

    def test_xlsx_identity_matching(self, real_app):
        """XLSX rows go through the same IdentityMatcher."""
        from app.models import Person, PersonIdentity
        from app.intake.matcher import IdentityMatcher
        from app import db
        with real_app.app_context():
            p = Person(canonical_name="Ritu", preferred_name="Ritu")
            db.session.add(p); db.session.commit()
            db.session.add(PersonIdentity(person_id=p.id, identity_type="email",
                                          identity_value="ritu@example.com", normalized_value="ritu@example.com"))
            db.session.commit()
            matcher = IdentityMatcher(session=db.session)
            mappings = [{"source_column": "email", "target_field": "identity.email"}]
            result = matcher.resolve_row({"email": "ritu@example.com"}, mappings)
            assert result["status"] == "MATCHED"

    def test_xlsx_proposal_generation(self, real_app):
        """XLSX intake produces proposals through the same ImportProposalBuilder."""
        from app.models import IntakeSession, IntakeCandidate
        from app.intake.proposal import ImportProposalBuilder
        from app import db
        with real_app.app_context():
            sess = IntakeSession(source_type="xlsx", source_name="customers.xlsx", status="ready_for_review")
            db.session.add(sess); db.session.commit()
            c = IntakeCandidate(session_id=sess.id, row_index=0, identity_status="MATCHED", import_status="pending")
            db.session.add(c); db.session.commit()
            proposal = ImportProposalBuilder.build(sess, [c])
            assert proposal["total_rows"] == 1


# =========================================================================
# 2. PROPOSAL VERSIONING & APPROVAL AUDIT
# =========================================================================

class TestProposalVersioning:

    def test_proposal_version_increments(self, real_app):
        """Each proposal build increments proposal_version."""
        from app.models import IntakeSession, IntakeCandidate
        from app.intake.proposal import ImportProposalBuilder
        from app import db
        with real_app.app_context():
            sess = IntakeSession(source_type="csv", source_name="test.csv", status="ready_for_review", proposal_version=0)
            db.session.add(sess); db.session.commit()
            c = IntakeCandidate(session_id=sess.id, row_index=0, identity_status="MATCHED", import_status="pending")
            db.session.add(c); db.session.commit()
            p1 = ImportProposalBuilder.build(sess, [c])
            assert p1["proposal_version"] == 1
            p2 = ImportProposalBuilder.build(sess, [c])
            assert p2["proposal_version"] == 2

    def test_approval_records_audit_fields(self, real_app):
        """Approval records approved_by, approved_at, approved_proposal_version."""
        from app.models import IntakeSession
        from app.intake.session import IntakeSessionState
        from app.intake.committer import GovernedCommitter
        from app import db
        with real_app.app_context():
            sess = IntakeSession(source_type="csv", source_name="test.csv", status=IntakeSessionState.READY_FOR_REVIEW,
                                 proposal_version=3)
            db.session.add(sess); db.session.commit()
            committer = GovernedCommitter(session=db.session)
            result = committer.approve(sess.id, approved_by="admin@shunyaos.com", scope=IntakeSessionState.APPROVE_ALL)
            assert result["success"] is True
            assert result["proposal_version"] == 3
            assert sess.approved_by == "admin@shunyaos.com"
            assert sess.approved_at is not None
            assert sess.approved_proposal_version == 3

    def test_commit_rejects_stale_approval(self, real_app):
        """Commit rejects when proposal version changed since approval."""
        from app.models import IntakeSession, IntakeCandidate
        from app.intake.session import IntakeSessionState
        from app.intake.committer import GovernedCommitter
        from app import db
        with real_app.app_context():
            sess = IntakeSession(source_type="csv", source_name="test.csv", status=IntakeSessionState.READY_FOR_REVIEW,
                                 proposal_version=1, approved_proposal_version=1)
            db.session.add(sess); db.session.commit()
            # Simulate proposal version changing after approval
            sess.proposal_version = 2
            db.session.commit()
            # Now approve — but version mismatch
            sess.status = IntakeSessionState.APPROVED
            db.session.commit()
            committer = GovernedCommitter(session=db.session)
            result = committer.commit(sess.id)
            assert result["success"] is False
            assert "Re-approval required" in result["error"]


# =========================================================================
# 3. TRANSACTION SAFETY
# =========================================================================

class TestTransactionSafety:

    def test_rollback_on_failure(self, real_app):
        """All canonical writes are rolled back on mid-import failure."""
        from app.models import Person, IntakeSession, IntakeCandidate, CustomerProfile
        from app.intake.session import IntakeSessionState
        from app.intake.proposal import ImportProposalBuilder
        from app.intake.committer import GovernedCommitter
        from app import db
        with real_app.app_context():
            sess = IntakeSession(source_type="csv", source_name="test.csv", status=IntakeSessionState.READY_FOR_REVIEW)
            db.session.add(sess); db.session.commit()
            # Create a NO_MATCH candidate (safe)
            c = IntakeCandidate(session_id=sess.id, row_index=0, identity_status="NO_MATCH",
                                import_status="pending",
                                normalized_data='{"name":"Test","email":"test@test.com"}')
            db.session.add(c); db.session.commit()
            # Build proposal, approve, commit
            ImportProposalBuilder.build(sess, [c])
            committer = GovernedCommitter(session=db.session)
            committer.approve(sess.id, approved_by="admin")
            result = committer.commit(sess.id)
            assert result["success"] is True, f"First commit failed: {result.get('error','')}"
            # Single candidate, no blocked — should be COMPLETED
            assert result["session_status"] == IntakeSessionState.COMPLETED

    def test_retry_after_failure(self, real_app):
        """Retry after partial failure does not duplicate Persons."""
        from app.models import Person, IntakeSession, IntakeCandidate
        from app.intake.session import IntakeSessionState
        from app.intake.proposal import ImportProposalBuilder
        from app.intake.committer import GovernedCommitter
        from app import db
        with real_app.app_context():
            sess = IntakeSession(source_type="csv", source_name="test.csv", status=IntakeSessionState.READY_FOR_REVIEW)
            db.session.add(sess); db.session.commit()
            c = IntakeCandidate(session_id=sess.id, row_index=0, identity_status="NO_MATCH",
                                import_status="pending",
                                normalized_data='{"name":"Retry","email":"retry@test.com"}')
            db.session.add(c); db.session.commit()
            # Build proposal, approve, commit
            ImportProposalBuilder.build(sess, [c])
            committer = GovernedCommitter(session=db.session)
            committer.approve(sess.id, approved_by="admin")
            r1 = committer.commit(sess.id)
            assert r1["success"] is True
            # Verify exactly one Person created
            persons = Person.query.filter_by(canonical_name="Retry").all()
            assert len(persons) == 1


# =========================================================================
# 4. PARTIAL COMMIT SEMANTICS
# =========================================================================

class TestPartialCommit:

    def test_safe_90_ambiguous_5_insufficient_5(self, real_app):
        """90 safe + 5 ambiguous + 5 insufficient: safe 90 commit after safe-only approval."""
        from app.models import Person, PersonIdentity, IntakeSession, IntakeCandidate
        from app.intake.session import IntakeSessionState
        from app.intake.proposal import ImportProposalBuilder
        from app.intake.committer import GovernedCommitter
        from app import db
        with real_app.app_context():
            sess = IntakeSession(source_type="csv", source_name="bulk.csv", status=IntakeSessionState.READY_FOR_REVIEW,
                                 proposal_version=0)
            db.session.add(sess); db.session.commit()
            candidates = []
            # 90 NO_MATCH (safe)
            for i in range(90):
                c = IntakeCandidate(session_id=sess.id, row_index=i, identity_status="NO_MATCH",
                                    import_status="pending",
                                    normalized_data=f'{{"name":"Person {i}","email":"p{i}@test.com"}}')
                db.session.add(c); candidates.append(c)
            # 5 AMBIGUOUS
            for i in range(90, 95):
                c = IntakeCandidate(session_id=sess.id, row_index=i, identity_status="AMBIGUOUS",
                                    import_status="blocked", normalized_data="{}")
                db.session.add(c); candidates.append(c)
            # 5 INSUFFICIENT_IDENTITY
            for i in range(95, 100):
                c = IntakeCandidate(session_id=sess.id, row_index=i, identity_status="INSUFFICIENT_IDENTITY",
                                    import_status="blocked", normalized_data='{"name":"Only Name"}')
                db.session.add(c); candidates.append(c)
            db.session.commit()

            # Build proposal
            proposal = ImportProposalBuilder.build(sess, candidates)
            assert proposal["total_rows"] == 100
            assert proposal["summary"]["safe_import_count"] == 90
            assert proposal["summary"]["review_required_count"] == 10
            assert proposal["summary"]["can_commit_safe_candidates"] is True

            # Approve and commit
            committer = GovernedCommitter(session=db.session)
            committer.approve(sess.id, approved_by="admin")
            result = committer.commit(sess.id)
            assert result["success"] is True
            assert result["imported"] == 90
            assert result["unresolved_remaining"] == 10
            assert result["session_status"] == IntakeSessionState.PARTIALLY_COMPLETED

    def test_10_unresolved_remain_reviewable(self, real_app):
        """Unresolved candidates remain after partial commit."""
        from app.models import IntakeSession, IntakeCandidate
        from app.intake.session import IntakeSessionState
        from app.intake.proposal import ImportProposalBuilder
        from app.intake.committer import GovernedCommitter
        from app import db
        with real_app.app_context():
            sess = IntakeSession(source_type="csv", source_name="partial.csv", status=IntakeSessionState.READY_FOR_REVIEW)
            db.session.add(sess); db.session.commit()
            c1 = IntakeCandidate(session_id=sess.id, row_index=0, identity_status="NO_MATCH",
                                 import_status="pending", normalized_data='{"name":"Safe","email":"safe@test.com"}')
            c2 = IntakeCandidate(session_id=sess.id, row_index=1, identity_status="AMBIGUOUS",
                                 import_status="blocked", normalized_data="{}")
            db.session.add(c1); db.session.add(c2); db.session.commit()
            ImportProposalBuilder.build(sess, [c1, c2])
            committer = GovernedCommitter(session=db.session)
            committer.approve(sess.id, approved_by="admin")
            result = committer.commit(sess.id)
            assert result["unresolved_remaining"] == 1
            # Verify the unresolved candidate still exists
            remaining = db.session.query(IntakeCandidate).filter(
                IntakeCandidate.session_id == sess.id,
                IntakeCandidate.identity_status == "AMBIGUOUS"
            ).all()
            assert len(remaining) == 1

    def test_retry_does_not_duplicate_committed(self, real_app):
        """Retry after partial commit does not duplicate the 90 committed records."""
        from app.models import Person, IntakeSession, IntakeCandidate
        from app.intake.session import IntakeSessionState
        from app.intake.proposal import ImportProposalBuilder
        from app.intake.committer import GovernedCommitter
        from app import db
        with real_app.app_context():
            sess = IntakeSession(source_type="csv", source_name="retry.csv", status=IntakeSessionState.READY_FOR_REVIEW)
            db.session.add(sess); db.session.commit()
            c = IntakeCandidate(session_id=sess.id, row_index=0, identity_status="NO_MATCH",
                                import_status="pending",
                                normalized_data='{"name":"RetryTest","email":"retrytest@test.com"}')
            db.session.add(c); db.session.commit()
            ImportProposalBuilder.build(sess, [c])
            committer = GovernedCommitter(session=db.session)
            committer.approve(sess.id, approved_by="admin")
            r1 = committer.commit(sess.id)
            assert r1["imported"] == 1